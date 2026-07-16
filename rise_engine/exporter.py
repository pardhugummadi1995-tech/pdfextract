"""Exporter: write the generated outputs as JSON, CSV and Excel."""

from __future__ import annotations

import csv
import dataclasses
import io
import json

from .model import ProjectModel

INDENT_HEADER = [
    "SKU Code",
    "Item Description",
    "Category",
    "Room(s)",
    "Cabinet(s)",
    "Total Qty",
    "UOM",
    "Source Page",
    "Flags",
]

REVIEW_HEADER = [
    "SKU Code",
    "Item Description",
    "Category",
    "Room(s)",
    "Cabinet(s)",
    "Extracted Qty",
    "UOM",
    "Source Page",
    "Flags",
    "Approved Qty",
    "Reviewer Notes",
]


def _indent_row(line):
    return [
        line.sku_code,
        line.description,
        line.category,
        ", ".join(line.rooms),
        ", ".join(line.cabinets),
        "" if line.total_qty is None else line.total_qty,
        line.uom or "",
        ", ".join(str(p) for p in line.pages),
        ", ".join(line.flags),
    ]


def to_dict(model: ProjectModel) -> dict:
    return dataclasses.asdict(model)


def to_json(model: ProjectModel) -> str:
    return json.dumps(to_dict(model), indent=2, ensure_ascii=False)


def indent_to_csv(model: ProjectModel) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerow(INDENT_HEADER)
    for line in model.material_indent:
        writer.writerow(_indent_row(line))
    return buf.getvalue()


def review_to_csv(model: ProjectModel) -> str:
    """Editable review sheet: an "Approved Qty" column is pre-filled with the
    extracted quantity (blank when it could not be read) for a human to confirm
    or correct before final export."""
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerow(REVIEW_HEADER)
    for line in model.material_indent:
        extracted = "" if line.total_qty is None else line.total_qty
        writer.writerow(
            [
                line.sku_code,
                line.description,
                line.category,
                ", ".join(line.rooms),
                ", ".join(line.cabinets),
                extracted,
                line.uom or "",
                ", ".join(str(p) for p in line.pages),
                ", ".join(line.flags),
                extracted,  # Approved Qty (edit me)
                "",  # Reviewer Notes
            ]
        )
    return buf.getvalue()


def write_review_csv(model: ProjectModel, path: str) -> None:
    with open(path, "w", encoding="utf-8", newline="") as fh:
        fh.write(review_to_csv(model))


def cabinets_to_csv(model: ProjectModel) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerow(["Cabinet", "Room", "Width", "Depth", "Height", "Carcass", "Shutter", "Page"])
    for c in model.cabinets:
        writer.writerow(
            [c.code, c.room, c.width or "", c.depth or "", c.height or "", c.carcass, c.shutter, c.page or ""]
        )
    return buf.getvalue()


def write_json(model: ProjectModel, path: str) -> None:
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(to_json(model))


def write_csv(model: ProjectModel, path: str) -> None:
    with open(path, "w", encoding="utf-8", newline="") as fh:
        fh.write(indent_to_csv(model))


def write_excel(model: ProjectModel, path: str) -> None:
    """Write a multi-sheet .xlsx workbook (requires openpyxl)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()

    ws = wb.active
    ws.title = "Material Indent"
    ws.append(INDENT_HEADER)
    for line in model.material_indent:
        ws.append(_indent_row(line))

    cab = wb.create_sheet("Cabinets")
    cab.append(["Cabinet", "Room", "Width", "Depth", "Height", "Carcass", "Shutter", "Page"])
    for c in model.cabinets:
        cab.append([c.code, c.room, c.width, c.depth, c.height, c.carcass, c.shutter, c.page])

    rs = wb.create_sheet("Room Summary")
    rs.append(
        [
            "Room",
            "No. of Cabinets",
            "Hardware Items",
            "Estimated Inventory",
            "Electrical Points",
            "Plumbing Points",
        ]
    )
    for r in model.room_summary:
        rs.append(
            [
                r.room,
                r.cabinets,
                r.hardware_lines,
                r.estimated_inventory,
                r.electrical_points,
                r.plumbing_points,
            ]
        )

    cc = wb.create_sheet("Category Counts")
    if model.category_counts:
        counts = model.category_counts
        cc.append(["Category", "Count"])
        for label, value in [
            ("Rooms", counts.rooms),
            ("Cabinets", counts.cabinets),
            ("Hardware Types", counts.hardware_types),
            ("Hardware Qty", counts.hardware_qty),
            ("Finish Codes", counts.finish_codes),
            ("Electrical Points", counts.electrical_points),
            ("Plumbing Points", counts.plumbing_points),
        ]:
            cc.append([label, value])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)

    wb.save(path)
